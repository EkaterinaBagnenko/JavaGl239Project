from msilib.schema import Font
import os
import sys
import time 

from openpyxl import *
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font

from screeninfo import get_monitors

from PyQt5 import QtGui, QtCore, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QLineEdit, QPushButton, QComboBox, QCheckBox, QScrollArea, QVBoxLayout, QWidget, QRadioButton, QTableWidgetItem, QHeaderView, QFileDialog, QScrollBar

from PyQt5.QtWidgets import QMessageBox, QTableWidget, QTableWidgetItem
from PyQt5.QtGui import QIcon, QFont, QPixmap
from PyQt5.QtCore import Qt, QSize

import shutil


class MainWindow(QMainWindow):                                                                                                  

    arrExc1 = []                                                                                                                                                                                                                              
    objects = []        

  #-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

    def baseSetup(self):                                                                                                        
        self.setWindowTitle('Neoclassica Structured Data')                                                                      
        self.setWindowIcon(QIcon('pictures/smalllogo.png'))                                                                     
        self.move(0, 0)                                                                                                         
        self.resize(1920, 1080)                                                                                                 
        self.setMinimumSize(1600, 900)                                                                                          
        self.setStyleSheet(" background-color: rgba(255, 255, 255, 255);")                                                      

    def __init__(self):                                                                                                         
        super().__init__()                                                                                                        
        self.baseSetup()                                                                                                        
        self.mainRoomUser()                                                                                                                                                                                                         

    def constantLogo(self):                                                                                                      
        self.smalllogo = QPixmap('pictures/smalllogo.png')
        self.small_logo = QLabel(self)
        self.small_logo.setPixmap(self.smalllogo)
        self.small_logo.setGeometry(1810, 920, 75, 75)

  #-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

    def mainRoomUser(self):                                                                                                         
        self.biglogo = QPixmap('pictures/biglogo.png')
        self.big_logo = QLabel(self)
        self.big_logo.setPixmap(self.biglogo)
        self.big_logo.setGeometry(320, 10, 1280, 480)
        self.big_logo.show()
        
        self.exc_archive = QPushButton("К сравнению таблиц за 2020/2021/2022", self)
        self.exc_archive.setGeometry(450, 300, 1020, 380)
        self.exc_archive.setFont(QtGui.QFont("Corbel", 40, QtGui.QFont.Bold))   
        self.exc_archive.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc_archive.clicked.connect(self.exc_archive_was_clicked)
        self.exc_archive.show()

        self.exc_rec1 = QPushButton("Архив всех таблиц по месяцам", self)
        self.exc_rec1.setGeometry(450, 700, 1020, 100)
        self.exc_rec1.setFont(QtGui.QFont("Corbel", 32, QtGui.QFont.Bold))   
        self.exc_rec1.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc_rec1.clicked.connect(self.exc_rec1_was_clicked)
        self.exc_rec1.show()

        self.exc_rec = QPushButton("Обновление данных", self)
        self.exc_rec.setGeometry(450, 820, 1020, 100)
        self.exc_rec.setFont(QtGui.QFont("Corbel", 32, QtGui.QFont.Bold))   
        self.exc_rec.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc_rec.clicked.connect(self.exc_rec_was_clicked)
        self.exc_rec.show()

        self.tableExc = QTableWidget(self)
        self.tableExc.hide()  
        self.table_Exc = QTableWidget(self)
        self.table_Exc.hide()  
        self.deskXL20 = QTableWidget(self)
        self.deskXL20.hide() 
        self.deskXL21 = QTableWidget(self)
        self.deskXL21.hide() 
        self.deskXL22 = QTableWidget(self)
        self.deskXL22.hide() 
  
    def XXmainRoomUser(self):                                                                                                     
        self.big_logo.hide()
        self.exc_archive.hide()
        self.exc_rec1.hide()
        self.exc_rec.hide()

    def goHome(self):                                                                                                            
        self.toHome = QPushButton(self)
        self.toHome.setGeometry(20, 20, 55, 55)
        self.toHome.setAutoFillBackground(True)    
        self.toHome.setIcon(QIcon('pictures/to_home.png'))  
        self.toHome.setIconSize(QSize(50, 50))
        self.toHome.setStyleSheet("border :2px solid black;" "border-radius: 10px;")
        self.toHome.clicked.connect(self.toHome_was_clicked)
        self.toHome.show()
                                                                                                                                                                 
    def toHome_was_clicked(self):                                                                                                   
        self.toHome.hide()

        for i in range (0, len(self.arrExc1)):
            self.arrExc1[i].hide()
        for i in range (0, len(self.objects)):
            self.objects[i].hide()

        self.tableExc.hide()
        self.table_Exc.hide()
        self.deskXL20.hide()
        self.deskXL21.hide()
        self.deskXL22.hide()

        self.mainRoomUser()

  #-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        
    def exc_rec1_was_clicked(self):
        self.XXmainRoomUser()
        self.goHome()

        self.exc_2020 = QPushButton("2020", self)
        self.exc_2021 = QPushButton("2021", self)
        self.exc_2022 = QPushButton("2022", self)
        self.excyear = QPushButton("ИТОГ", self)

        btnhh = 160
        btnww = 60 
        k = 280
        self.exc_2022.setGeometry(610 - k, 120, btnhh, btnww)
        self.exc_2022.setAutoFillBackground(True)    
        self.exc_2022.setFont(QtGui.QFont("Corbel", 16, QtGui.QFont.Bold))   
        self.exc_2021.setGeometry(790 - k, 120, btnhh, btnww)
        self.exc_2021.setAutoFillBackground(True)    
        self.exc_2021.setFont(QtGui.QFont("Corbel", 16, QtGui.QFont.Bold))   
        self.exc_2020.setGeometry(970 - k, 120, btnhh, btnww)
        self.exc_2020.setAutoFillBackground(True)    
        self.exc_2020.setFont(QtGui.QFont("Corbel", 16, QtGui.QFont.Bold))   
        self.excyear.setGeometry(1150 - k, 120, btnhh, btnww)
        self.excyear.setAutoFillBackground(True)    
        self.excyear.setFont(QtGui.QFont("Corbel", 16, QtGui.QFont.Bold))   
       
        self.exc_2020.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc_2021.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc_2022.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.excyear.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")

        self.exc_2020.clicked.connect(self.exc_2020_was_clicked)
        self.exc_2021.clicked.connect(self.exc_2021_was_clicked)
        self.exc_2022.clicked.connect(self.exc_2022_was_clicked)
        self.excyear.clicked.connect(self.excyear_was_clicked)

        self.arrExc1.append(self.exc_2020)
        self.exc_2020.show()
        self.arrExc1.append(self.exc_2021)
        self.exc_2021.show()
        self.arrExc1.append(self.exc_2022)
        self.exc_2022.show()
        self.arrExc1.append(self.excyear)
        self.excyear.show()

        self.tableExc.setColumnCount(2) 
        self.tableExc.setRowCount(12)  
        self.tableExc.setHorizontalHeaderLabels(["Месяц и год", "Наличие таблицы"])
        self.tableExc.horizontalHeaderItem(0).setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
        self.tableExc.horizontalHeaderItem(1).setFont(QFont('Corbel', 12, QtGui.QFont.Bold))   
        self.tableExc.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) 
        self.tableExc.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) 
    
        www = 350
        hhh = 50
        self.tableExc.setColumnWidth(0, www)
        self.tableExc.setColumnWidth(1, www)

        for i in range(0, 13):
            self.tableExc.setRowHeight(i, hhh)

        self.tableExc.setGeometry(590 - k, 250, 740, 660)
        self.tableExc.show()

        self.open_file = QPushButton("Открыть файл в Excel", self)
        self.open_file.setGeometry(1250, 510, 360, 80)
        self.open_file.setAutoFillBackground(True)    
        self.open_file.setFont(QtGui.QFont("Corbel", 24, QtGui.QFont.Bold))
        self.open_file.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}"
         "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
         "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")

        self.texxt = "notthis"
        self.open_file.clicked.connect(self.open_file_was_clicked)
        self.objects.append(self.open_file)
 
    def exc_2020_was_clicked(self): 
        self.fill_the_table_2020()
        self.hardfill_the_table_2020()
        self.exc_2020.setStyleSheet("background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);" "border-radius: 10px;")
        self.exc_2021.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc_2022.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.excyear.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.open_file.hide()

    def exc_2021_was_clicked(self):
        self.fill_the_table_2021()
        self.hardfill_the_table_2021()
        self.exc_2020.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc_2021.setStyleSheet("background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);" "border-radius: 10px;")
        self.exc_2022.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.excyear.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.open_file.hide()

    def exc_2022_was_clicked(self):       
        self.fill_the_table_2022()
        self.hardfill_the_table_2022()
        self.exc_2020.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc_2021.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc_2022.setStyleSheet("background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);" "border-radius: 10px;" ) 
        self.excyear.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.open_file.hide()

    def excyear_was_clicked(self):
        self.fill_the_tableyear()
        self.hardfill_the_tableyear()

        self.excyear.setStyleSheet("background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);" "border-radius: 10px;")
        self.exc_2021.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc_2022.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc_2020.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")

    def fill_the_table_2020(self):
        newItem = QTableWidgetItem ('Январь 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(0,0, newItem)

        newItem = QTableWidgetItem ('Февраль 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(1,0, newItem)

        newItem = QTableWidgetItem ('Март 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(2,0, newItem)

        newItem = QTableWidgetItem ('Апрель 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(3,0, newItem)

        newItem = QTableWidgetItem ('Май 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(4,0, newItem)

        newItem = QTableWidgetItem ('Июнь 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(5,0, newItem)

        newItem = QTableWidgetItem ('Июль 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(6,0, newItem)

        newItem = QTableWidgetItem ('Август 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(7,0, newItem)

        newItem = QTableWidgetItem ('Сентябрь 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(8,0, newItem)

        newItem = QTableWidgetItem ('Октябрь 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(9,0, newItem)

        newItem = QTableWidgetItem ('Ноябрь 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(10,0, newItem)

        newItem = QTableWidgetItem ('Декабрь 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(11,0, newItem)

    def fill_the_table_2021(self):
        newItem = QTableWidgetItem ('Январь 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(0,0, newItem)

        newItem = QTableWidgetItem ('Февраль 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(1,0, newItem)

        newItem = QTableWidgetItem ('Март 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(2,0, newItem)

        newItem = QTableWidgetItem ('Апрель 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(3,0, newItem)

        newItem = QTableWidgetItem ('Май 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(4,0, newItem)

        newItem = QTableWidgetItem ('Июнь 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(5,0, newItem)

        newItem = QTableWidgetItem ('Июль 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(6,0, newItem)

        newItem = QTableWidgetItem ('Август 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(7,0, newItem)

        newItem = QTableWidgetItem ('Сентябрь 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(8,0, newItem)

        newItem = QTableWidgetItem ('Октябрь 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(9,0, newItem)

        newItem = QTableWidgetItem ('Ноябрь 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(10,0, newItem)

        newItem = QTableWidgetItem ('Декабрь 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(11,0, newItem)

    def fill_the_table_2022(self):          
        newItem = QTableWidgetItem ('Январь 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(0,0, newItem)

        newItem = QTableWidgetItem ('Февраль 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(1,0, newItem)

        newItem = QTableWidgetItem ('Март 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(2,0, newItem)

        newItem = QTableWidgetItem ('Апрель 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(3,0, newItem)

        newItem = QTableWidgetItem ('Май 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(4,0, newItem)

        newItem = QTableWidgetItem ('Июнь 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(5,0, newItem)

        newItem = QTableWidgetItem ('Июль 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(6,0, newItem)

        newItem = QTableWidgetItem ('Август 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(7,0, newItem)

        newItem = QTableWidgetItem ('Сентябрь 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(8,0, newItem)

        newItem = QTableWidgetItem ('Октябрь 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(9,0, newItem)

        newItem = QTableWidgetItem ('Ноябрь 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(10,0, newItem)

        newItem = QTableWidgetItem ('Декабрь 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(11,0, newItem)

    def fill_the_tableyear(self):
        newItem = QTableWidgetItem ('Итоговый 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(0,0, newItem)

        newItem = QTableWidgetItem ('Итоговый 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(1,0, newItem)

        newItem = QTableWidgetItem ('Итоговый 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(2,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(3,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(4,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(5,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(6,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(7,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(8,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(9,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(10,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.tableExc.setItem(11,0, newItem)

    def hardfill_the_table_2020(self):
        self.jan20 = QPushButton(self)
        self.tableExc.setCellWidget(0, 1, self.jan20)
        self.feb20 = QPushButton(self)
        self.tableExc.setCellWidget(1, 1, self.feb20)
        self.mar20 = QPushButton(self)
        self.tableExc.setCellWidget(2, 1, self.mar20)
        self.apr20 = QPushButton(self)
        self.tableExc.setCellWidget(3, 1, self.apr20)
        self.may20 = QPushButton(self)
        self.tableExc.setCellWidget(4, 1, self.may20)
        self.jun20 = QPushButton(self)
        self.tableExc.setCellWidget(5, 1, self.jun20)
        self.jul20 = QPushButton(self)
        self.tableExc.setCellWidget(6, 1, self.jul20)
        self.aug20 = QPushButton(self)
        self.tableExc.setCellWidget(7, 1, self.aug20)
        self.sep20 = QPushButton(self)
        self.tableExc.setCellWidget(8, 1, self.sep20)
        self.oct20 = QPushButton(self)
        self.tableExc.setCellWidget(9, 1, self.oct20)
        self.nov20 = QPushButton(self)
        self.tableExc.setCellWidget(10, 1, self.nov20)
        self.dec20 = QPushButton(self)
        self.tableExc.setCellWidget(11, 1, self.dec20)

        arr20 =[]
        arr20.append(self.jan20)
        arr20.append(self.feb20)
        arr20.append(self.mar20)
        arr20.append(self.apr20)
        arr20.append(self.may20)
        arr20.append(self.jun20)
        arr20.append(self.jul20)
        arr20.append(self.aug20)
        arr20.append(self.sep20)
        arr20.append(self.oct20)
        arr20.append(self.nov20)
        arr20.append(self.dec20)

        arr20n =[]
        arr20n.append("jan20.xlsx")
        arr20n.append("feb20.xlsx")
        arr20n.append("mar20.xlsx")
        arr20n.append("apr20.xlsx")
        arr20n.append("may20.xlsx")
        arr20n.append("jun20.xlsx")
        arr20n.append("jul20.xlsx")
        arr20n.append("aug20.xlsx")
        arr20n.append("sep20.xlsx")
        arr20n.append("oct20.xlsx")
        arr20n.append("nov20.xlsx")
        arr20n.append("dec20.xlsx")

        for i in range (0, 12):
            strpath = 'DATA/EXC/archive/2020/' + str(arr20n[i])
            if(os.path.exists(strpath)):
                arr20[i].setText(arr20n[i])
                arr20[i].setIcon(QIcon('pictures/excel.png'))  
                arr20[i].setIconSize(QSize(20, 20))
                arr20[i].setFont(QtGui.QFont("Corbel", 12)) 
                arr20[i].clicked.connect(self.excParcer0)
            else:
                arr20[i].setText("---")
                arr20[i].clicked.connect(self.excTrouble)

    def hardfill_the_table_2021(self):
        self.jan21 = QPushButton(self)
        self.tableExc.setCellWidget(0, 1, self.jan21)
        self.feb21 = QPushButton(self)
        self.tableExc.setCellWidget(1, 1, self.feb21)
        self.mar21 = QPushButton(self)
        self.tableExc.setCellWidget(2, 1, self.mar21)
        self.apr21 = QPushButton(self)
        self.tableExc.setCellWidget(3, 1, self.apr21)
        self.may21 = QPushButton(self)
        self.tableExc.setCellWidget(4, 1, self.may21)
        self.jun21 = QPushButton(self)
        self.tableExc.setCellWidget(5, 1, self.jun21)
        self.jul21 = QPushButton(self)
        self.tableExc.setCellWidget(6, 1, self.jul21)
        self.aug21 = QPushButton(self)
        self.tableExc.setCellWidget(7, 1, self.aug21)
        self.sep21 = QPushButton(self)
        self.tableExc.setCellWidget(8, 1, self.sep21)
        self.oct21 = QPushButton(self)
        self.tableExc.setCellWidget(9, 1, self.oct21)
        self.nov21 = QPushButton(self)
        self.tableExc.setCellWidget(10, 1, self.nov21)
        self.dec21 = QPushButton(self)
        self.tableExc.setCellWidget(11, 1, self.dec21)

        arr21 =[]
        arr21.append(self.jan21)
        arr21.append(self.feb21)
        arr21.append(self.mar21)
        arr21.append(self.apr21)
        arr21.append(self.may21)
        arr21.append(self.jun21)
        arr21.append(self.jul21)
        arr21.append(self.aug21)
        arr21.append(self.sep21)
        arr21.append(self.oct21)
        arr21.append(self.nov21)
        arr21.append(self.dec21)

        arr21n =[]
        arr21n.append("jan21.xlsx")
        arr21n.append("feb21.xlsx")
        arr21n.append("mar21.xlsx")
        arr21n.append("apr21.xlsx")
        arr21n.append("may21.xlsx")
        arr21n.append("jun21.xlsx")
        arr21n.append("jul21.xlsx")
        arr21n.append("aug21.xlsx")
        arr21n.append("sep21.xlsx")
        arr21n.append("oct21.xlsx")
        arr21n.append("nov21.xlsx")
        arr21n.append("dec21.xlsx")

        for i in range (0, 12):
            strpath = 'DATA/EXC/archive/2021/' + str(arr21n[i])
            if(os.path.exists(strpath)):
                arr21[i].setText(arr21n[i])
                arr21[i].setIcon(QIcon('pictures/excel.png'))  
                arr21[i].setIconSize(QSize(21, 21))
                arr21[i].setFont(QtGui.QFont("Corbel", 12)) 
                arr21[i].clicked.connect(self.excParcer0)

            else:
                arr21[i].setText("---")
                arr21[i].clicked.connect(self.excTrouble)

    def hardfill_the_table_2022(self):     
        self.jan22 = QPushButton(self)
        self.tableExc.setCellWidget(0, 1, self.jan22)
        self.feb22 = QPushButton(self)
        self.tableExc.setCellWidget(1, 1, self.feb22)
        self.mar22 = QPushButton(self)
        self.tableExc.setCellWidget(2, 1, self.mar22)
        self.apr22 = QPushButton(self)
        self.tableExc.setCellWidget(3, 1, self.apr22)
        self.may22 = QPushButton(self)
        self.tableExc.setCellWidget(4, 1, self.may22)
        self.jun22 = QPushButton(self)
        self.tableExc.setCellWidget(5, 1, self.jun22)
        self.jul22 = QPushButton(self)
        self.tableExc.setCellWidget(6, 1, self.jul22)
        self.aug22 = QPushButton(self)
        self.tableExc.setCellWidget(7, 1, self.aug22)
        self.sep22 = QPushButton(self)
        self.tableExc.setCellWidget(8, 1, self.sep22)
        self.oct22 = QPushButton(self)
        self.tableExc.setCellWidget(9, 1, self.oct22)
        self.nov22 = QPushButton(self)
        self.tableExc.setCellWidget(10, 1, self.nov22)
        self.dec22 = QPushButton(self)
        self.tableExc.setCellWidget(11, 1, self.dec22)

        arr22 =[]
        arr22.append(self.jan22)
        arr22.append(self.feb22)
        arr22.append(self.mar22)
        arr22.append(self.apr22)
        arr22.append(self.may22)
        arr22.append(self.jun22)
        arr22.append(self.jul22)
        arr22.append(self.aug22)
        arr22.append(self.sep22)
        arr22.append(self.oct22)
        arr22.append(self.nov22)
        arr22.append(self.dec22)

        arr22n =[]
        arr22n.append("jan22.xlsx")
        arr22n.append("feb22.xlsx")
        arr22n.append("mar22.xlsx")
        arr22n.append("apr22.xlsx")
        arr22n.append("may22.xlsx")
        arr22n.append("jun22.xlsx")
        arr22n.append("jul22.xlsx")
        arr22n.append("aug22.xlsx")
        arr22n.append("sep22.xlsx")
        arr22n.append("oct22.xlsx")
        arr22n.append("nov22.xlsx")
        arr22n.append("dec22.xlsx")

        for i in range (0, 12):
            strpath = 'DATA/EXC/archive/2022/' + str(arr22n[i])
            if(os.path.exists(strpath)):
                arr22[i].setText(arr22n[i])
                arr22[i].setIcon(QIcon('pictures/excel.png'))  
                arr22[i].setIconSize(QSize(22, 22))
                arr22[i].setFont(QtGui.QFont("Corbel", 12)) 
                arr22[i].clicked.connect(self.excParcer0)

            else:
                arr22[i].setText("---")
                arr22[i].clicked.connect(self.excTrouble)

    def hardfill_the_tableyear(self):
        self.y_2020 = QPushButton(self)
        self.tableExc.setCellWidget(0, 1, self.y_2020)
        self.y_2021 = QPushButton(self)
        self.tableExc.setCellWidget(1, 1, self.y_2021)
        self.y_2022 = QPushButton(self)
        self.tableExc.setCellWidget(2, 1, self.y_2022)

        arry = []
        arry.append(self.y_2020)
        arry.append(self.y_2021)
        arry.append(self.y_2022)
        
        arryn =[]
        arryn.append("2020.xlsx")
        arryn.append("2021.xlsx")
        arryn.append("2022.xlsx")

        for i in range (0, 3):
            strpath = 'DATA/EXC/archive/' + str(arryn[i])
            if(os.path.exists(strpath)):
                arry[i].setText(arryn[i])
                arry[i].setIcon(QIcon('pictures/excel.png'))  
                arry[i].setIconSize(QSize(22, 22))
                arry[i].setFont(QtGui.QFont("Corbel", 12)) 
                arry[i].clicked.connect(self.excParcer0)

            else:
                arry[i].setText("---")
                arry[i].clicked.connect(self.excTrouble) 

        for i in range (3, 12):
            self.tableExc.removeCellWidget(i, 1)

    def excTrouble(self):   
        self.open_file.hide()                                                                                                       
        self.exc_tr = QMessageBox(self)
        self.exc_tr.setWindowTitle("Инфо")
        self.exc_tr.move(800, 450)
        self.exc_tr.setText("Данный файл ещё не был загружен")
        self.exc_tr.setIcon(QMessageBox.Icon.Information)
        self.exc_tr.exec() 
     
    def open_file_was_clicked(self):
        name = self.texxt
        try:
            strbase = "DATA/EXC/archive/"
            name1 = '20' + name[3:5]
            file = strbase + name1 + "/" + name
            filefull = os.path.abspath(file)
            os.startfile(filefull)
        except:
            strbase = "DATA/EXC/archive"
            file = strbase + "/" + name
            filefull = os.path.abspath(file)
            os.startfile(filefull)

    def excParcer0(self):                  
        self.open_file.show()  
        texxt1 = self.sender()
        self.texxt = texxt1.text()

  #-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

    def exc_rec_was_clicked(self):
        self.XXmainRoomUser()
        self.goHome()
        self.name = ""

        self.exc__2020 = QPushButton("2020", self)
        self.exc__2021 = QPushButton("2021", self)
        self.exc__2022 = QPushButton("2022", self)
        self.exc__year = QPushButton("ИТОГ", self)

        btnhh = 160
        btnww = 60 
        k = 280
        self.exc__2022.setGeometry(610 - k, 120, btnhh, btnww)
        self.exc__2022.setAutoFillBackground(True)    
        self.exc__2022.setFont(QtGui.QFont("Corbel", 16, QtGui.QFont.Bold))   
        self.exc__2021.setGeometry(790 - k, 120, btnhh, btnww)
        self.exc__2021.setAutoFillBackground(True)    
        self.exc__2021.setFont(QtGui.QFont("Corbel", 16, QtGui.QFont.Bold))   
        self.exc__2020.setGeometry(970 - k, 120, btnhh, btnww)
        self.exc__2020.setAutoFillBackground(True)    
        self.exc__2020.setFont(QtGui.QFont("Corbel", 16, QtGui.QFont.Bold))   
        self.exc__year.setGeometry(1150 - k, 120, btnhh, btnww)
        self.exc__year.setAutoFillBackground(True)    
        self.exc__year.setFont(QtGui.QFont("Corbel", 16, QtGui.QFont.Bold))   

        self.exc__2020.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc__2021.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc__2022.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc__year.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")

        self.exc__2020.clicked.connect(self.exc__2020_was_clicked)
        self.exc__2021.clicked.connect(self.exc__2021_was_clicked)
        self.exc__2022.clicked.connect(self.exc__2022_was_clicked)
        self.exc__year.clicked.connect(self.exc__year_was_clicked)

        self.arrExc1.append(self.exc__2020)
        self.exc__2020.show()
        self.arrExc1.append(self.exc__2021)
        self.exc__2021.show()
        self.arrExc1.append(self.exc__2022)
        self.exc__2022.show()
        self.arrExc1.append(self.exc__year)
        self.exc__year.show()

        self.table_Exc.setColumnCount(2) 
        self.table_Exc.setRowCount(12)  
        self.table_Exc.setHorizontalHeaderLabels(["Месяц / год", "Наличие таблицы"])
        self.table_Exc.horizontalHeaderItem(0).setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
        self.table_Exc.horizontalHeaderItem(1).setFont(QFont('Corbel', 12, QtGui.QFont.Bold))   
        self.table_Exc.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) 
        self.table_Exc.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) 
    
        www = 350
        hhh = 50
        self.table_Exc.setColumnWidth(0, www)
        self.table_Exc.setColumnWidth(1, www)
        
        for i in range(0, 13):
            self.table_Exc.setRowHeight(i, hhh)

        self.table_Exc.setGeometry(590 - k, 250, 740, 660)
        self.table_Exc.show()
        
        self.dobav = QPushButton("Добавить файл", self)
        self.dobav.setGeometry(1250, 510, 360, 80)
        self.dobav.setAutoFillBackground(True)    
        self.dobav.setFont(QtGui.QFont("Corbel", 24, QtGui.QFont.Bold))
        self.dobav.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}"
         "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
         "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.dobav.clicked.connect(self.dobav_was_clicked)
        
        self.obnov = QPushButton("Обновить файл", self)
        self.obnov.setGeometry(1250, 510, 360, 80)
        self.obnov.setAutoFillBackground(True)    
        self.obnov.setFont(QtGui.QFont("Corbel", 24, QtGui.QFont.Bold))
        self.obnov.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}"
         "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
         "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.obnov.clicked.connect(self.obnov_was_clicked)

        self.arrExc1.append(self.dobav)
        self.arrExc1.append(self.obnov)

    def exc__2020_was_clicked(self):
        self.fill_the_table__2020()
        self.hardfill_the_table__2020()
        self.obnov.hide()
        self.dobav.hide()

        self.exc__2020.setStyleSheet("background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);" "border-radius: 10px;")
        self.exc__2021.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc__2022.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc__year.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")

    def exc__2021_was_clicked(self):
        self.fill_the_table__2021()
        self.hardfill_the_table__2021()
        self.obnov.hide()
        self.dobav.hide()

        self.exc__2021.setStyleSheet("background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);" "border-radius: 10px;")
        self.exc__2020.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc__2022.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc__year.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")

    def exc__2022_was_clicked(self):
        self.fill_the_table__2022()
        self.hardfill_the_table__2022()
        self.obnov.hide()
        self.dobav.hide()

        self.exc__2022.setStyleSheet("background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);" "border-radius: 10px;")
        self.exc__2021.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc__2020.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc__year.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")

    def exc__year_was_clicked(self):
        self.fill_the_table_year()
        self.hardfill_the_table__year()
        self.obnov.hide()
        self.dobav.hide()

        self.exc__year.setStyleSheet("background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);" "border-radius: 10px;")
        self.exc__2021.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc__2022.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.exc__2020.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")

    def fill_the_table_year(self):
        newItem = QTableWidgetItem ('Итоговый 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(0,0, newItem)

        newItem = QTableWidgetItem ('Итоговый 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(1,0, newItem)

        newItem = QTableWidgetItem ('Итоговый 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(2,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(3,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(4,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(5,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(6,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(7,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(8,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(9,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(10,0, newItem)

        newItem = QTableWidgetItem ('')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(11,0, newItem)

    def fill_the_table__2020(self):
        newItem = QTableWidgetItem ('Январь 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(0,0, newItem)

        newItem = QTableWidgetItem ('Февраль 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(1,0, newItem)

        newItem = QTableWidgetItem ('Март 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(2,0, newItem)

        newItem = QTableWidgetItem ('Апрель 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(3,0, newItem)

        newItem = QTableWidgetItem ('Май 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(4,0, newItem)

        newItem = QTableWidgetItem ('Июнь 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(5,0, newItem)

        newItem = QTableWidgetItem ('Июль 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(6,0, newItem)

        newItem = QTableWidgetItem ('Август 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(7,0, newItem)

        newItem = QTableWidgetItem ('Сентябрь 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(8,0, newItem)

        newItem = QTableWidgetItem ('Октябрь 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(9,0, newItem)

        newItem = QTableWidgetItem ('Ноябрь 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(10,0, newItem)

        newItem = QTableWidgetItem ('Декабрь 2020')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(11,0, newItem)

    def fill_the_table__2021(self):
        newItem = QTableWidgetItem ('Январь 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(0,0, newItem)

        newItem = QTableWidgetItem ('Февраль 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(1,0, newItem)

        newItem = QTableWidgetItem ('Март 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(2,0, newItem)

        newItem = QTableWidgetItem ('Апрель 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(3,0, newItem)

        newItem = QTableWidgetItem ('Май 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(4,0, newItem)

        newItem = QTableWidgetItem ('Июнь 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(5,0, newItem)

        newItem = QTableWidgetItem ('Июль 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(6,0, newItem)

        newItem = QTableWidgetItem ('Август 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(7,0, newItem)

        newItem = QTableWidgetItem ('Сентябрь 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(8,0, newItem)

        newItem = QTableWidgetItem ('Октябрь 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(9,0, newItem)

        newItem = QTableWidgetItem ('Ноябрь 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(10,0, newItem)

        newItem = QTableWidgetItem ('Декабрь 2021')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(11,0, newItem)
    
    def fill_the_table__2022(self):
        newItem = QTableWidgetItem ('Январь 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(0,0, newItem)

        newItem = QTableWidgetItem ('Февраль 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(1,0, newItem)

        newItem = QTableWidgetItem ('Март 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(2,0, newItem)

        newItem = QTableWidgetItem ('Апрель 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(3,0, newItem)

        newItem = QTableWidgetItem ('Май 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(4,0, newItem)

        newItem = QTableWidgetItem ('Июнь 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(5,0, newItem)

        newItem = QTableWidgetItem ('Июль 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(6,0, newItem)

        newItem = QTableWidgetItem ('Август 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(7,0, newItem)

        newItem = QTableWidgetItem ('Сентябрь 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(8,0, newItem)

        newItem = QTableWidgetItem ('Октябрь 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(9,0, newItem)

        newItem = QTableWidgetItem ('Ноябрь 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(10,0, newItem)

        newItem = QTableWidgetItem ('Декабрь 2022')
        newItem.setTextAlignment(Qt.AlignCenter)
        newItem.setFont(QFont('Corbel', 12))
        self.table_Exc.setItem(11,0, newItem)

    def hardfill_the_table__2020(self):
        arr_20n =[]
        arr_20n.append("jan20.xlsx")
        arr_20n.append("feb20.xlsx")
        arr_20n.append("mar20.xlsx")
        arr_20n.append("apr20.xlsx")
        arr_20n.append("may20.xlsx")
        arr_20n.append("jun20.xlsx")
        arr_20n.append("jul20.xlsx")
        arr_20n.append("aug20.xlsx")
        arr_20n.append("sep20.xlsx")
        arr_20n.append("oct20.xlsx")
        arr_20n.append("nov20.xlsx")
        arr_20n.append("dec20.xlsx")

        self.jan_20 = QPushButton(self, objectName = arr_20n[0])
        self.table_Exc.setCellWidget(0, 1, self.jan_20)
        self.feb_20 = QPushButton(self, objectName = arr_20n[1])
        self.table_Exc.setCellWidget(1, 1, self.feb_20)
        self.mar_20 = QPushButton(self, objectName = arr_20n[2])
        self.table_Exc.setCellWidget(2, 1, self.mar_20)
        self.apr_20 = QPushButton(self, objectName = arr_20n[3])
        self.table_Exc.setCellWidget(3, 1, self.apr_20)
        self.may_20 = QPushButton(self, objectName = arr_20n[4])
        self.table_Exc.setCellWidget(4, 1, self.may_20)
        self.jun_20 = QPushButton(self, objectName = arr_20n[5])
        self.table_Exc.setCellWidget(5, 1, self.jun_20)
        self.jul_20 = QPushButton(self, objectName = arr_20n[6])
        self.table_Exc.setCellWidget(6, 1, self.jul_20)
        self.aug_20 = QPushButton(self, objectName = arr_20n[7])
        self.table_Exc.setCellWidget(7, 1, self.aug_20)
        self.sep_20 = QPushButton(self, objectName = arr_20n[8])
        self.table_Exc.setCellWidget(8, 1, self.sep_20)
        self.oct_20 = QPushButton(self, objectName = arr_20n[9])
        self.table_Exc.setCellWidget(9, 1, self.oct_20)
        self.nov_20 = QPushButton(self, objectName = arr_20n[10])
        self.table_Exc.setCellWidget(10, 1, self.nov_20)
        self.dec_20 = QPushButton(self, objectName = arr_20n[11])
        self.table_Exc.setCellWidget(11, 1, self.dec_20)

        arr_20 = []
        arr_20.append(self.jan_20)
        arr_20.append(self.feb_20)
        arr_20.append(self.mar_20)
        arr_20.append(self.apr_20)
        arr_20.append(self.may_20)
        arr_20.append(self.jun_20)
        arr_20.append(self.jul_20)
        arr_20.append(self.aug_20)
        arr_20.append(self.sep_20)
        arr_20.append(self.oct_20)
        arr_20.append(self.nov_20)
        arr_20.append(self.dec_20)

        for i in range (0, 12):
            strpath = 'DATA/EXC/archive/2020/' + str(arr_20n[i])
            if(os.path.exists(strpath)):
                arr_20[i].setText(arr_20n[i])
                arr_20[i].setIcon(QIcon('pictures/excel.png'))  
                arr_20[i].setIconSize(QSize(20, 20))
                arr_20[i].setFont(QtGui.QFont("Corbel", 12)) 
                arr_20[i].clicked.connect(self.obn_and_dob)
            else:
                arr_20[i].setText("---")
                arr_20[i].clicked.connect(self.dob)

    def hardfill_the_table__2021(self):
        arr_21n =[]
        arr_21n.append("jan21.xlsx")
        arr_21n.append("feb21.xlsx")
        arr_21n.append("mar21.xlsx")
        arr_21n.append("apr21.xlsx")
        arr_21n.append("may21.xlsx")
        arr_21n.append("jun21.xlsx")
        arr_21n.append("jul21.xlsx")
        arr_21n.append("aug21.xlsx")
        arr_21n.append("sep21.xlsx")
        arr_21n.append("oct21.xlsx")
        arr_21n.append("nov21.xlsx")
        arr_21n.append("dec21.xlsx")

        self.jan_21 = QPushButton(self, objectName = arr_21n[0])
        self.table_Exc.setCellWidget(0, 1, self.jan_21)
        self.feb_21 = QPushButton(self, objectName = arr_21n[1])
        self.table_Exc.setCellWidget(1, 1, self.feb_21)
        self.mar_21 = QPushButton(self, objectName = arr_21n[2])
        self.table_Exc.setCellWidget(2, 1, self.mar_21)
        self.apr_21 = QPushButton(self, objectName = arr_21n[3])
        self.table_Exc.setCellWidget(3, 1, self.apr_21)
        self.may_21 = QPushButton(self, objectName = arr_21n[4])
        self.table_Exc.setCellWidget(4, 1, self.may_21)
        self.jun_21 = QPushButton(self, objectName = arr_21n[5])
        self.table_Exc.setCellWidget(5, 1, self.jun_21)
        self.jul_21 = QPushButton(self, objectName = arr_21n[6])
        self.table_Exc.setCellWidget(6, 1, self.jul_21)
        self.aug_21 = QPushButton(self, objectName = arr_21n[7])
        self.table_Exc.setCellWidget(7, 1, self.aug_21)
        self.sep_21 = QPushButton(self, objectName = arr_21n[8])
        self.table_Exc.setCellWidget(8, 1, self.sep_21)
        self.oct_21 = QPushButton(self, objectName = arr_21n[9])
        self.table_Exc.setCellWidget(9, 1, self.oct_21)
        self.nov_21 = QPushButton(self, objectName = arr_21n[10])
        self.table_Exc.setCellWidget(10, 1, self.nov_21)
        self.dec_21 = QPushButton(self, objectName = arr_21n[11])
        self.table_Exc.setCellWidget(11, 1, self.dec_21)

        arr_21 =[]
        arr_21.append(self.jan_21)
        arr_21.append(self.feb_21)
        arr_21.append(self.mar_21)
        arr_21.append(self.apr_21)
        arr_21.append(self.may_21)
        arr_21.append(self.jun_21)
        arr_21.append(self.jul_21)
        arr_21.append(self.aug_21)
        arr_21.append(self.sep_21)
        arr_21.append(self.oct_21)
        arr_21.append(self.nov_21)
        arr_21.append(self.dec_21)


        for i in range (0, 12):
            strpath = 'DATA/EXC/archive/2021/' + str(arr_21n[i])
            if(os.path.exists(strpath)):
                arr_21[i].setText(arr_21n[i])
                arr_21[i].setIcon(QIcon('pictures/excel.png'))  
                arr_21[i].setIconSize(QSize(21, 21))
                arr_21[i].setFont(QtGui.QFont("Corbel", 12)) 
                arr_21[i].clicked.connect(self.obn_and_dob)

            else:
                arr_21[i].setText("---")
                arr_21[i].clicked.connect(self.dob)

    def hardfill_the_table__2022(self):     
        arr_22n =[]
        arr_22n.append("jan22.xlsx")
        arr_22n.append("feb22.xlsx")
        arr_22n.append("mar22.xlsx")
        arr_22n.append("apr22.xlsx")
        arr_22n.append("may22.xlsx")
        arr_22n.append("jun22.xlsx")
        arr_22n.append("jul22.xlsx")
        arr_22n.append("aug22.xlsx")
        arr_22n.append("sep22.xlsx")
        arr_22n.append("oct22.xlsx")
        arr_22n.append("nov22.xlsx")
        arr_22n.append("dec22.xlsx")

        self.jan_22 = QPushButton(self, objectName = arr_22n[0])
        self.table_Exc.setCellWidget(0, 1, self.jan_22)
        self.feb_22 = QPushButton(self, objectName = arr_22n[1])
        self.table_Exc.setCellWidget(1, 1, self.feb_22)
        self.mar_22 = QPushButton(self, objectName = arr_22n[2])
        self.table_Exc.setCellWidget(2, 1, self.mar_22)
        self.apr_22 = QPushButton(self, objectName = arr_22n[3])
        self.table_Exc.setCellWidget(3, 1, self.apr_22)
        self.may_22 = QPushButton(self, objectName = arr_22n[4])
        self.table_Exc.setCellWidget(4, 1, self.may_22)
        self.jun_22 = QPushButton(self, objectName = arr_22n[5])
        self.table_Exc.setCellWidget(5, 1, self.jun_22)
        self.jul_22 = QPushButton(self, objectName = arr_22n[6])
        self.table_Exc.setCellWidget(6, 1, self.jul_22)
        self.aug_22 = QPushButton(self, objectName = arr_22n[7])
        self.table_Exc.setCellWidget(7, 1, self.aug_22)
        self.sep_22 = QPushButton(self, objectName = arr_22n[8])
        self.table_Exc.setCellWidget(8, 1, self.sep_22)
        self.oct_22 = QPushButton(self, objectName = arr_22n[9])
        self.table_Exc.setCellWidget(9, 1, self.oct_22)
        self.nov_22 = QPushButton(self, objectName = arr_22n[10])
        self.table_Exc.setCellWidget(10, 1, self.nov_22)
        self.dec_22 = QPushButton(self, objectName = arr_22n[11])
        self.table_Exc.setCellWidget(11, 1, self.dec_22)

        arr_22 =[]
        arr_22.append(self.jan_22)
        arr_22.append(self.feb_22)
        arr_22.append(self.mar_22)
        arr_22.append(self.apr_22)
        arr_22.append(self.may_22)
        arr_22.append(self.jun_22)
        arr_22.append(self.jul_22)
        arr_22.append(self.aug_22)
        arr_22.append(self.sep_22)
        arr_22.append(self.oct_22)
        arr_22.append(self.nov_22)
        arr_22.append(self.dec_22)

        for i in range (0, 12):
            strpath = 'DATA/EXC/archive/2022/' + str(arr_22n[i])
            if(os.path.exists(strpath)):
                arr_22[i].setText(arr_22n[i])
                arr_22[i].setIcon(QIcon('pictures/excel.png'))  
                arr_22[i].setIconSize(QSize(22, 22))
                arr_22[i].setFont(QtGui.QFont("Corbel", 12)) 
                arr_22[i].clicked.connect(self.obn_and_dob)

            else:
                arr_22[i].setText("---")
                arr_22[i].clicked.connect(self.dob) 
    
    def hardfill_the_table__year(self):
        arr_yn =[]
        arr_yn.append("2020.xlsx")
        arr_yn.append("2021.xlsx")
        arr_yn.append("2022.xlsx")

        self.y2020 = QPushButton(self, objectName = arr_yn[0])
        self.table_Exc.setCellWidget(0, 1, self.y2020)
        self.y2021 = QPushButton(self, objectName = arr_yn[1])
        self.table_Exc.setCellWidget(1, 1, self.y2021)
        self.y2022 = QPushButton(self, objectName = arr_yn[2])
        self.table_Exc.setCellWidget(2, 1, self.y2022)

        arr_y = []
        arr_y.append(self.y2020)
        arr_y.append(self.y2021)
        arr_y.append(self.y2022) 
        
        for i in range (0, 3):
            strpath = 'DATA/EXC/archive/' + str(arr_yn[i])
            if(os.path.exists(strpath)):
                arr_y[i].setText(arr_yn[i])
                arr_y[i].setIcon(QIcon('pictures/excel.png'))  
                arr_y[i].setIconSize(QSize(22, 22))
                arr_y[i].setFont(QtGui.QFont("Corbel", 12)) 
                arr_y[i].clicked.connect(self.obn_and_dob)

            else:
                arr_y[i].setText("---")
                arr_y[i].clicked.connect(self.dob) 

        for i in range (3, 12):
            self.table_Exc.removeCellWidget(i, 1)

    def obn_and_dob(self):
        self.obnov.show()
        self.dobav.hide()
        button = QApplication.instance().sender()
        self.name = button.objectName()

    def dob(self):
        self.obnov.hide()
        self.dobav.show()
        button = QApplication.instance().sender()
        self.name = button.objectName()

    def dobav_was_clicked(self):
        path = QFileDialog.getOpenFileName(self, filter="Excel (*.xlsx )")
        if(path[0] != ""):
            ppp = path[0]
            first= self.name[0:3]
            second = self.name[3:5]

            if(first.isdigit() == True):
                second = self.name[2:4]
                second = int(second)

                if(second == 20):
                    path = shutil.copy(ppp, 'DATA/EXC/archive/2020.xlsx')
                if(second == 21):
                    path = shutil.copy(ppp, 'DATA/EXC/archive/2021.xlsx')
                if(second == 22):
                    path = shutil.copy(ppp, 'DATA/EXC/archive/2022.xlsx')

                self.exc__year_was_clicked()        
            else:
                second = int(second)
               
                if(second == 20):
                    direct =  'DATA/EXC/archive/2020'
                    direct = direct + "/" + self.name
                    path = shutil.copy(ppp, direct)
                    self.exc__2020_was_clicked()    
                if(second == 21):
                    direct =  'DATA/EXC/archive/2021'
                    direct = direct + "/" + self.name
                    path = shutil.copy(ppp, direct)
                    self.exc__2021_was_clicked()    
                if(second == 22):
                    direct =  'DATA/EXC/archive/2022'
                    direct = direct + "/" + self.name
                    path = shutil.copy(ppp, direct)
                    self.exc__2022_was_clicked()    

    def obnov_was_clicked(self):
        path = QFileDialog.getOpenFileName(self, filter="Excel (*.xlsx )")
        if(path[0] != ""):
            ppp = path[0]
            first= self.name[0:3]
            second = self.name[3:5]

            if(first.isdigit() == True):
                second = self.name[2:4]
                second = int(second)
                
                if(second == 20):
                    os.remove('DATA/EXC/archive/2020.xlsx')
                    path = shutil.copy(ppp, 'DATA/EXC/archive/2020.xlsx')
                if(second == 21):
                    os.remove('DATA/EXC/archive/2021.xlsx')
                    path = shutil.copy(ppp, 'DATA/EXC/archive/2021.xlsx')
                if(second == 22):
                    os.remove('DATA/EXC/archive/2022.xlsx')
                    path = shutil.copy(ppp, 'DATA/EXC/archive/2022.xlsx')

                self.exc__year_was_clicked()        
            else:
                second = int(second)
               
                if(second == 20):
                    direct =  'DATA/EXC/archive/2020'
                    direct = direct + "/" + self.name
                    os.remove(direct)
                    path = shutil.copy(ppp, direct)
                    self.exc__2020_was_clicked()    
                if(second == 21):
                    direct =  'DATA/EXC/archive/2021'
                    direct = direct + "/" + self.name
                    os.remove(direct)
                    path = shutil.copy(ppp, direct)
                    self.exc__2021_was_clicked()    
                if(second == 22):
                    direct =  'DATA/EXC/archive/2022'
                    direct = direct + "/" + self.name
                    os.remove(direct)
                    path = shutil.copy(ppp, direct)
                    self.exc__2022_was_clicked()    

  #-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

    def exc_archive_was_clicked(self):                                                                                          
        self.XXmainRoomUser()
        self.goHome()

        self.start2task()

    def start2task(self):
        self.arrKnopok = []

        self.choose2020 = QCheckBox("2020", self)
        self.choose2020.setFont(QtGui.QFont("Corbel", 18))
        self.choose2020.setGeometry(180, 20, 120, 30)
        self.choose2020.setCheckState(Qt.CheckState.Checked)
        self.choose2020.show()
        self.arrKnopok.append(self.choose2020)
        self.objects.append(self.choose2020)

        self.choose2021 = QCheckBox("2021", self)
        self.choose2021.setFont(QtGui.QFont("Corbel", 18))
        self.choose2021.setGeometry(180, 70, 120, 30)
        self.choose2021.setCheckState(Qt.CheckState.Checked)
        self.choose2021.show()
        self.arrKnopok.append(self.choose2021)
        self.objects.append(self.choose2021)

        self.choose2022 = QCheckBox("2022", self)
        self.choose2022.setFont(QtGui.QFont("Corbel", 18))
        self.choose2022.setGeometry(180, 120, 120, 30)
        self.choose2022.setCheckState(Qt.CheckState.Checked)
        self.choose2022.show()
        self.arrKnopok.append(self.choose2022)
        self.objects.append(self.choose2022)

        self.ramka = QLabel(self)
        self.ramka.setStyleSheet("border :1px solid black;")
        self.ramka.setGeometry(10, 180, 1910, 2)
        self.ramka.show()
        self.arrExc1.append(self.ramka)

        self.lay = QVBoxLayout(self)
        self.arrKnopok2 = []

        self.c1 = QRadioButton("Ежедневная прибыль с каждого", self)
        self.c1.setFont(QtGui.QFont("Corbel", 16))
        self.c1.show()
        self.arrKnopok2.append(self.c1)
        self.objects.append(self.c1)
        self.lay.addWidget(self.c1)

        self.c2 = QRadioButton("Ежедневные другие доходы", self)
        self.c2.setFont(QtGui.QFont("Corbel", 16))
        self.c2.show()
        self.arrKnopok2.append(self.c2)
        self.objects.append(self.c2)
        self.lay.addWidget(self.c2)

        self.widget = QWidget(self)
        self.widget.setLayout(self.lay)

        self.scroll = QScrollArea(self)
        self.scroll.setGeometry(320, 15, 700, 150)
        self.scroll.setAutoFillBackground(True)
        self.scroll.setWidget(self.widget) 
        self.scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.scroll.show()
        self.arrExc1.append(self.scroll)

        self.sbros = QPushButton("Сброс", self)
        self.sbros.setAutoFillBackground(True)    
        self.sbros.setFont(QtGui.QFont("Corbel", 18, QtGui.QFont.Bold))   
        self.sbros.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.sbros.clicked.connect(self.sbros_was_clicked)
        self.sbros.setGeometry(1100, 40, 200, 100)
        self.arrExc1.append(self.sbros)

        self.primenit = QPushButton("Применить", self)
        self.primenit.setAutoFillBackground(True)    
        self.primenit.setFont(QtGui.QFont("Corbel", 18, QtGui.QFont.Bold))   
        self.primenit.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.primenit.clicked.connect(self.primenit_was_clicked)
        self.primenit.setGeometry(1320, 40, 200, 100)
        self.primenit.show()
        self.arrExc1.append(self.primenit)

        self.arrKnopok2[0].setChecked(True)

        self.wb2020 = load_workbook("DATA/EXC/archive/2020.xlsx")
        self.ss20 = self.wb2020.active
        self.wb2021 = load_workbook("DATA/EXC/archive/2021.xlsx")
        self.ss21 = self.wb2021.active
        self.wb2022 = load_workbook("DATA/EXC/archive/2022.xlsx")
        self.ss22 = self.wb2022.active

        self.deskXLZ()

        self.save1 = QPushButton("Сохранить в файл", self)
        self.save1.setAutoFillBackground(True)    
        self.save1.setFont(QtGui.QFont("Corbel", 18, QtGui.QFont.Bold))   
        self.save1.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.save1.clicked.connect(self.save1_was_clicked)
        self.save1.setGeometry(1550, 40, 300, 100)
        self.arrExc1.append(self.save1)

        self.save2 = QPushButton("Сохранить в файл", self)
        self.save2.setAutoFillBackground(True)    
        self.save2.setFont(QtGui.QFont("Corbel", 18, QtGui.QFont.Bold))   
        self.save2.setStyleSheet("QPushButton { border :2px solid black;" "border-radius: 10px;" "background-color: rgba(202, 163, 86, 50);}" 
        "QPushButton:hover { background-color: rgba(202, 163, 86, 255); border-color: rgba(255, 255, 255, 255); color: rgba(255, 255, 255, 255);}"
        "QPushButton:pressed { background-color: rgba(202, 163, 86, 255); border-color: rgba(0, 0, 0, 255); color: rgba(0, 0, 0, 255);}")
        self.save2.clicked.connect(self.save1_was_clicked)
        self.save2.setGeometry(1550, 40, 300, 100)
        self.arrExc1.append(self.save2)

        self.verSum = QScrollBar(self)
        self.verSum.setStyleSheet("background-color: rgba(202, 163, 86, 255)")

        self.deskXL20.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.deskXL20.setVerticalScrollBar(self.verSum)
        self.deskXL21.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.deskXL21.setVerticalScrollBar(self.verSum)
        self.deskXL22.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.deskXL22.setVerticalScrollBar(self.verSum)

        self.bar = QScrollArea(self)
        self.bar.setVerticalScrollBar(self.verSum)
        self.bar.setGeometry(1850, 200, 20, 750)
        self.arrExc1.append(self.bar)
        self.arrExc1.append(self.verSum)

        self.horSum = QScrollBar(self)
        self.horSum.setStyleSheet("background-color: rgba(202, 163, 86, 255)")

        self.deskXL20.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.deskXL20.setHorizontalScrollBar(self.horSum)
        self.deskXL21.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.deskXL21.setHorizontalScrollBar(self.horSum)
        self.deskXL22.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.deskXL22.setHorizontalScrollBar(self.horSum)

        self.bar2 = QScrollArea(self)
        self.bar2.setHorizontalScrollBar(self.horSum)
        self.bar2.setGeometry(110, 980, 1650, 20)
        self.arrExc1.append(self.bar2)
        self.arrExc1.append(self.horSum)

    def deskXLZ(self):

        self.deskXL20.setColumnCount(53)
        self.deskXL20.setRowCount(2)
        self.deskXL21.setColumnCount(53)
        self.deskXL21.setRowCount(2)
        self.deskXL22.setColumnCount(53)
        self.deskXL22.setRowCount(2)

        self.deskXL20.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) 
        self.deskXL20.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) 
        self.deskXL21.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) 
        self.deskXL21.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) 
        self.deskXL22.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) 
        self.deskXL22.verticalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed) 
    
        www = 55
        hhh = 30  

        for i in range(1, 53):
            self.deskXL20.setColumnWidth(i, www)
            self.deskXL20.setRowHeight(i, hhh)
            self.deskXL21.setColumnWidth(i, www)
            self.deskXL21.setRowHeight(i, hhh)
            self.deskXL22.setColumnWidth(i, www)
            self.deskXL22.setRowHeight(i, hhh)

        self.deskXL20.setColumnWidth(0, 10)
        self.deskXL21.setColumnWidth(0, 10)
        self.deskXL22.setColumnWidth(0, 10)
        
        self.deskXL20.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.deskXL20.horizontalHeader().setVisible(False)
        self.deskXL21.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.deskXL21.horizontalHeader().setVisible(False)
        self.deskXL22.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.deskXL22.horizontalHeader().setVisible(False)

        t = 1
        while(t < 52):
            self.deskXL20.setSpan(0, t, 1, 4)
            self.deskXL21.setSpan(0, t, 1, 4)
            self.deskXL22.setSpan(0, t, 1, 4)

            if(t == 1):
                item = QTableWidgetItem('Январь 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Январь 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Январь 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t == 5):
                item = "Февраль 2020"
                item = QTableWidgetItem(item)
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Февраль 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Февраль 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t == 9):
                item = QTableWidgetItem('Март 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Март 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Март 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t == 13):
                item = QTableWidgetItem('Апрель 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Апрель 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Апрель 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t == 17):
                item = QTableWidgetItem('Май 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Май 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Май 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t == 21):
                item = QTableWidgetItem('Июнь 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Июнь 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Июнь 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t == 25):
                item = QTableWidgetItem('Июль 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Июль 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Июль 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t == 29):
                item = QTableWidgetItem('Август 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Август 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Август 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t == 33):
                item = QTableWidgetItem('Сентябрь 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Сентябрь 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Сентябрь 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t == 37):
                item = QTableWidgetItem('Октябрь 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Октябрь 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Октябрь 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t == 41):
                item = QTableWidgetItem('Ноябрь 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Ноябрь 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Ноябрь 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t == 45):
                item = QTableWidgetItem('Декабрь 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Декабрь 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Декабрь 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t == 49):
                item = QTableWidgetItem('ВЕСЬ 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('ВЕСЬ 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('ВЕСЬ 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            t = t + 4

        obj = QTableWidgetItem('ФИО')
        obj.setTextAlignment(Qt.AlignCenter)
        obj.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
        self.deskXL20.setVerticalHeaderItem(1, obj)
        obj2 = QTableWidgetItem('ФИО')
        obj2.setTextAlignment(Qt.AlignCenter)
        obj2.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
        self.deskXL21.setVerticalHeaderItem(1, obj2)
        obj3 = QTableWidgetItem('ФИО')
        obj3.setTextAlignment(Qt.AlignCenter)
        obj3.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
        self.deskXL22.setVerticalHeaderItem(1, obj3)

        obj = QTableWidgetItem('')
        obj.setTextAlignment(Qt.AlignCenter)
        obj.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
        self.deskXL20.setVerticalHeaderItem(0, obj)
        obj2 = QTableWidgetItem('')
        obj2.setTextAlignment(Qt.AlignCenter)
        obj2.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
        self.deskXL21.setVerticalHeaderItem(0, obj2)
        obj3 = QTableWidgetItem('')
        obj3.setTextAlignment(Qt.AlignCenter)
        obj3.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
        self.deskXL22.setVerticalHeaderItem(0, obj3)

        t = 1
        while(t < 53):
            if(t % 4 == 1):
                item = QTableWidgetItem('Нал бел')
                self.deskXL20.setItem(1, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 10, QtGui.QFont.Bold))
                item = QTableWidgetItem('Нал бел')
                self.deskXL21.setItem(1, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 10, QtGui.QFont.Bold))
                item = QTableWidgetItem('Нал бел')
                self.deskXL22.setItem(1, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 10, QtGui.QFont.Bold))
            if(t % 4 == 2):
                item = QTableWidgetItem('Нал сер')
                self.deskXL20.setItem(1, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 10, QtGui.QFont.Bold))
                item = QTableWidgetItem('Нал сер')
                self.deskXL21.setItem(1, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 10, QtGui.QFont.Bold))
                item = QTableWidgetItem('Нал сер')
                self.deskXL22.setItem(1, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 10, QtGui.QFont.Bold))
            if(t % 4 == 3):
                item = QTableWidgetItem('Безнал')
                self.deskXL20.setItem(1, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 10, QtGui.QFont.Bold))
                item = QTableWidgetItem('Безнал')
                self.deskXL21.setItem(1, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 10, QtGui.QFont.Bold))
                item = QTableWidgetItem('Безнал')
                self.deskXL22.setItem(1, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 10, QtGui.QFont.Bold))
            if(t % 4 == 0):
                item = QTableWidgetItem('ИТОГО')
                self.deskXL20.setItem(1, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 10, QtGui.QFont.Bold))
                item = QTableWidgetItem('ИТОГО')
                self.deskXL21.setItem(1, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 10, QtGui.QFont.Bold))
                item = QTableWidgetItem('ИТОГО')
                self.deskXL22.setItem(1, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 10, QtGui.QFont.Bold))
            t = t + 1
        
        t = 1
        while(t < 14):
            if(t % 13 == 1):
                item = QTableWidgetItem('Январь 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Январь 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Январь 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
            
            if(t % 13 == 2):
                item = "Февраль 2020"
                item = QTableWidgetItem(item)
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Февраль 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Февраль 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))

            if(t % 13 == 3):
                item = QTableWidgetItem('Март 2020')
                self.deskXL20.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Март 2021')
                self.deskXL21.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))
                item = QTableWidgetItem('Март 2022')
                self.deskXL22.setItem(0, t, item)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Corbel', 12, QtGui.QFont.Bold))

            t = t + 1    
        
    def save1_was_clicked(self):
        path = QFileDialog.getExistingDirectory(self)

        t = time.localtime()
        ct = time.strftime("%d%m%Y_%H%M%S", t)
        ct = ct + ".xlsx"
        wbnew = Workbook()
        ws = wbnew.active
        arr = [[0] * 53]
        arr.pop(0)

        arr1 = []
        ws.column_dimensions['A'].width = 40

        kk = 0
        height = 0

        if(self.arrKnopok[0].isChecked() == True):
            aa = 2
            bb = 5
            k = 9
            for i in range(1, 14):
                a = get_column_letter(aa)
                b = get_column_letter(bb)
                ss = a + "1:" + b + "1"
                ws.merge_cells(ss)
                sss = a + "1"
            
                if(i <= 3):
                    if(self.deskXL20.item(0, i) is None):  
                        ws[sss] = None
                    else:
                        ws[sss] = self.deskXL20.item(0, i).text()
                else:
                    if(self.deskXL20.item(0, i + k) is None):  
                        ws[sss] = None
                    else:
                        ws[sss] = self.deskXL20.item(0, i + k).text()
                    k = k + 3
                ws[sss].alignment = Alignment(horizontal='center')
                ws[sss].font = Font(size = 12, bold = True)
                aa = aa + 4
                bb = bb + 4
            height = height + 1

            aa = 1
            k = 2
            for i in range(0, 53):
                a = get_column_letter(aa)
                sss = a + str(k)
                if(self.deskXL20.item(1, i) is None):  
                        ws[sss] = None
                else:
                    ws[sss] = self.deskXL20.item(1, i).text()
                ws[sss].alignment = Alignment(horizontal='center')
                ws[sss].font = Font(size = 11, bold = True)
                aa = aa + 1
            height = height + 1

            aa = 1
            kk = 3
            for i in range(2, self.deskXL20.rowCount()):
                for j in range(0, 53):
                    if(j == 0):
                        item = self.deskXL20.verticalHeaderItem(i)
                    else:
                        item = self.deskXL20.item(i, j)

                    a = get_column_letter(aa)
                    ss = a + str(kk)

                    if(item is None):  
                        ws[ss] = None
                    else:
                        try:
                            ws[ss] = int(item.text())
                        except:
                            ws[ss] = item.text()

                    aa = aa + 1
                    ws[ss].alignment = Alignment(horizontal='center')

                aa = 1
                kk = kk + 1
                height = height + 1

            height = height + 2

        if(self.arrKnopok[1].isChecked() == True):
            aa = 2
            bb = 5
            k = 9
            for i in range(1, 14):
                a = get_column_letter(aa)
                b = get_column_letter(bb)
                ss = a + str(1 + height) + ":" + b + str(1 + height)
                ws.merge_cells(ss)
                sss = a + str(1 + height)
                if(i <= 3):
                    if(self.deskXL21.item(0, i) is None):  
                        ws[sss] = None
                    else:
                        ws[sss] = self.deskXL21.item(0, i).text()
                else:
                    if(self.deskXL21.item(0, i + k) is None):  
                        ws[sss] = None
                    else:
                        ws[sss] = self.deskXL21.item(0, i + k).text()
                    k = k + 3
                ws[sss].alignment = Alignment(horizontal='center')
                ws[sss].font = Font(size = 12, bold = True)
                aa = aa + 4
                bb = bb + 4

            aa = 1
            k = 2 + height
            for i in range(0, 53):
                a = get_column_letter(aa)
                sss = a + str(k)
                if(self.deskXL21.item(1, i) is None):  
                        ws[sss] = None
                else:
                    ws[sss] = self.deskXL21.item(1, i).text()

                ws[sss].alignment = Alignment(horizontal='center')
                ws[sss].font = Font(size = 11, bold = True)
                aa = aa + 1

            aa = 1
            kk = height + 3
            for i in range(2, self.deskXL21.rowCount()):
                for j in range(0, 53):
                    if(j == 0):
                        item = self.deskXL21.verticalHeaderItem(i)
                    else:
                        item = self.deskXL21.item(i, j)

                    a = get_column_letter(aa)
                    ss = a + str(kk)

                    if(item is None):  
                        ws[ss] = None
                    else:
                        try:
                            ws[ss] = int(item.text())
                        except:
                            ws[ss] = item.text()

                    aa = aa + 1
                    ws[ss].alignment = Alignment(horizontal='center')
                aa = 1
                kk = kk + 1
                height = height + 1

            height = height + 4
        
        if(self.arrKnopok[2].isChecked() == True):
            aa = 2
            bb = 5
            k = 9
            for i in range(1, 14):
                a = get_column_letter(aa)
                b = get_column_letter(bb)
                ss = a + str(1 + height) + ":" + b + str(1 + height)
                ws.merge_cells(ss)
                sss = a + str(1 + height)
            
                if(i <= 3):
                    if(self.deskXL22.item(0, i) is None):  
                        ws[sss] = None
                    else:
                        ws[sss] = self.deskXL22.item(0, i).text()
                else:
                    if(self.deskXL22.item(0, i + k) is None):  
                        ws[sss] = None
                    else:
                        ws[sss] = self.deskXL22.item(0, i + k).text()
                    k = k + 3
                ws[sss].alignment = Alignment(horizontal='center')
                ws[sss].font = Font(size = 12, bold = True)
                aa = aa + 4
                bb = bb + 4

            aa = 1
            k = 2 + height
            for i in range(0, 53):
                a = get_column_letter(aa)
                sss = a + str(k)
                if(self.deskXL22.item(1, i) is None):  
                        ws[sss] = None
                else:
                    ws[sss] = self.deskXL22.item(1, i).text()
                ws[sss].alignment = Alignment(horizontal='center')
                ws[sss].font = Font(size = 11, bold = True)
                aa = aa + 1

            aa = 1
            kk = height + 3
            for i in range(2, self.deskXL22.rowCount()):
                for j in range(0, 53):
                    if(j == 0):
                        item = self.deskXL22.verticalHeaderItem(i)
                    else:
                        item = self.deskXL22.item(i, j)

                    a = get_column_letter(aa)
                    ss = a + str(kk)

                    if(item is None):  
                        ws[ss] = None
                    else:
                        try:
                            ws[ss] = int(item.text())
                        except:
                            ws[ss] = item.text()
                    aa = aa + 1
                    ws[ss].alignment = Alignment(horizontal='center')
                aa = 1
                kk = kk + 1
                height = height + 1

            height = height + 2

        path = path + "/" + ct
        wbnew.save(path)

    def sbros_was_clicked(self):
        for i in range (0, len(self.arrKnopok)):
            self.arrKnopok[i].setCheckState(Qt.CheckState.Checked)
        
        self.arrKnopok2[0].setChecked(True)
        self.vbar = self.scroll.verticalScrollBar()
        self.vbar.setValue(self.vbar.minimum())
        
        self.deskXL20.hide()
        self.deskXL21.hide()
        self.deskXL22.hide()

        self.deskXLZ()
        
        self.save1.hide()
        self.sbros.hide()
        self.primenit.show()

        self.bar.hide()
        self.bar2.hide()
        self.verSum.hide()
        self.horSum.hide()

    def primenit_was_clicked(self):
        self.deskXL20.hide()
        self.deskXL21.hide()
        self.deskXL22.hide()

        self.primenit.hide()
        self.sbros.show()

        self.k = 0
        self.deskXLZ()
        self.hh = 250

        for i in range (0, len(self.arrKnopok)):
            if(self.arrKnopok[i].isChecked() == True):
                self.k = self.k + 1 
        if(self.k == 3):
            self.hh = 250
        if(self.k == 2):
            self.hh = 350
        if(self.k == 1):
            self.hh = 500

        if(self.k > 0):

            if(self.c1.isChecked() == True):
                self.ttt1()
                self.save1.show()

            if(self.c2.isChecked() == True):
                self.ttt2()
                self.save1.show()
            

        else:
            self.k_tr = QMessageBox(self)
            self.k_tr.setWindowTitle("Проблема")
            self.k_tr.move(800, 450)
            self.k_tr.setText("Выберите хотя бы один год")
            self.k_tr.setIcon(QMessageBox.Icon.Information)
            self.k_tr.exec() 
        
    def ttt1(self):
        nstr = 0

        if(self.arrKnopok[0].isChecked() == True):
            sss = 'B7'
            k_sss = 7

            while((self.ss20[sss].value) != 'ПРОЧИЕ Доходы'):
                ob = self.ss20[sss].value    
                if(not(ob is None)):
                    ob = ob.strip(' ') 
                    if(ob != 'ТЕРАПИЯ' and ob != 'ОРТОДОНТИЯ' and ob != 'ХИРУРГИЯ' and ob != 'КОСМЕТОЛОГИЯ'):

                        rowPosition = self.deskXL20.rowCount()
                        self.deskXL20.insertRow(rowPosition)

                        item = QTableWidgetItem(ob)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFont(QFont('Corbel', 12))
                        self.deskXL20.setVerticalHeaderItem(rowPosition, item)
                        ck = column_index_from_string('D')
                        ck1 = get_column_letter(ck)
                        ckstr = ck1 + str(k_sss)
                        summ = 0
                        sn = 0
                        sm = 0
                        sb = 0

                        while(ck < column_index_from_string('AZ')):

                            item = self.ss20[ckstr].value
                            if(item is None):
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                                summ = summ + 0
                            else:
                                if(ck % 4 != 3):
                                    summ = summ + item
                                    if(ck % 4 == 0):
                                        sn = sn + item
                                    if(ck % 4 == 1):
                                        sm = sm + item
                                    if(ck % 4 == 2):
                                        sb = sb + item
                                    
                                else:
                                    item = summ
                                    summ = 0
                                
                                item = str(item)
                                item = QTableWidgetItem(item)
                                self.deskXL20.setItem(rowPosition, ck - 3, QTableWidgetItem(item))
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                            ckstr = ck1 + str(k_sss)

                        item = QTableWidgetItem(str(sn))
                        self.deskXL20.setItem(rowPosition, ck - 3, item)
                        item = QTableWidgetItem(str(sm))
                        self.deskXL20.setItem(rowPosition, ck - 2, item)
                        item = QTableWidgetItem(str(sb))
                        self.deskXL20.setItem(rowPosition, ck - 1, item)
                        item = QTableWidgetItem(str(sn + sb + sm))
                        self.deskXL20.setItem(rowPosition, ck, item)

                        sn = 0
                        sb = 0
                        sm = 0


                        nstr = nstr + 1
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)

                    else:
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)
                else:
                    k_sss = k_sss + 1
                    sss = sss[0] + str(k_sss)
            
            self.deskXL20.setGeometry(100, 190, 1670, self.hh)
            self.deskXL20.show()

        if(self.arrKnopok[1].isChecked() == True):
            sss = 'B7'
            k_sss = 7
            while((self.ss21[sss].value) != 'ПРОЧИЕ Доходы'):
                ob = self.ss21[sss].value    
                if(not(ob is None)):
                    ob = ob.strip(' ') 
                    if(ob != 'ТЕРАПИЯ' and ob != 'ОРТОДОНТИЯ' and ob != 'ХИРУРГИЯ' and ob != 'КОСМЕТОЛОГИЯ'):

                        rowPosition = self.deskXL21.rowCount()
                        self.deskXL21.insertRow(rowPosition)

                        item = QTableWidgetItem(ob)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFont(QFont('Corbel', 12))
                        self.deskXL21.setVerticalHeaderItem(rowPosition, item)
                        ck = column_index_from_string('D')
                        ck1 = get_column_letter(ck)
                        ckstr = ck1 + str(k_sss)
                        summ = 0
                        sn = 0
                        sm = 0
                        sb = 0

                        while(ck < column_index_from_string('AZ')):

                            item = self.ss21[ckstr].value
                            if(item is None):
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                                summ = summ + 0
                            else:
                                if(ck % 4 != 3):
                                    summ = summ + item
                                    if(ck % 4 == 0):
                                        sn = sn + item
                                    if(ck % 4 == 1):
                                        sm = sm + item
                                    if(ck % 4 == 2):
                                        sb = sb + item
                                    
                                else:
                                    item = summ
                                    summ = 0
                                
                                item = str(item)
                                item = QTableWidgetItem(item)
                                self.deskXL21.setItem(rowPosition, ck - 3, QTableWidgetItem(item))
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                            ckstr = ck1 + str(k_sss)

                        item = QTableWidgetItem(str(sn))
                        self.deskXL21.setItem(rowPosition, ck - 3, item)
                        item = QTableWidgetItem(str(sm))
                        self.deskXL21.setItem(rowPosition, ck - 2, item)
                        item = QTableWidgetItem(str(sb))
                        self.deskXL21.setItem(rowPosition, ck - 1, item)
                        item = QTableWidgetItem(str(sn + sb + sm))
                        self.deskXL21.setItem(rowPosition, ck, item)

                        sn = 0
                        sb = 0
                        sm = 0


                        nstr = nstr + 1
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)

                    else:
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)
                else:
                    k_sss = k_sss + 1
                    sss = sss[0] + str(k_sss)
            
            if(self.k == 1):
                self.deskXL21.setGeometry(450, 190, 1370, self.hh)
            else:
                if(self.arrKnopok[0].isChecked() == True):
                    self.deskXL21.setGeometry(100, 190 + 10 + self.hh, 1670, self.hh)
                else:
                    self.deskXL21.setGeometry(100, 190, 1670, self.hh)
            self.deskXL21.show()

        if(self.arrKnopok[2].isChecked() == True):
            sss = 'B7'
            k_sss = 7
            while((self.ss22[sss].value) != 'ПРОЧИЕ Доходы'):
                ob = self.ss22[sss].value    
                if(not(ob is None)):
                    ob = ob.strip(' ') 
                    if(ob != 'ТЕРАПИЯ' and ob != 'ОРТОДОНТИЯ' and ob != 'ХИРУРГИЯ' and ob != 'КОСМЕТОЛОГИЯ'):

                        rowPosition = self.deskXL22.rowCount()
                        self.deskXL22.insertRow(rowPosition)

                        item = QTableWidgetItem(ob)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFont(QFont('Corbel', 12))
                        self.deskXL22.setVerticalHeaderItem(rowPosition, item)
                        ck = column_index_from_string('D')
                        ck1 = get_column_letter(ck)
                        ckstr = ck1 + str(k_sss)
                        summ = 0
                        sn = 0
                        sm = 0
                        sb = 0

                        while(ck < column_index_from_string('AZ')):

                            item = self.ss22[ckstr].value
                            if(item is None):
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                                summ = summ + 0
                            else:
                                if(ck % 4 != 3):
                                    summ = summ + item
                                    if(ck % 4 == 0):
                                        sn = sn + item
                                    if(ck % 4 == 1):
                                        sm = sm + item
                                    if(ck % 4 == 2):
                                        sb = sb + item
                                    
                                else:
                                    item = summ
                                    summ = 0
                                
                                item = str(item)
                                item = QTableWidgetItem(item)
                                self.deskXL22.setItem(rowPosition, ck - 3, QTableWidgetItem(item))
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                            ckstr = ck1 + str(k_sss)

                        item = QTableWidgetItem(str(sn))
                        self.deskXL22.setItem(rowPosition, ck - 3, item)
                        item = QTableWidgetItem(str(sm))
                        self.deskXL22.setItem(rowPosition, ck - 2, item)
                        item = QTableWidgetItem(str(sb))
                        self.deskXL22.setItem(rowPosition, ck - 1, item)
                        item = QTableWidgetItem(str(sn + sb + sm))
                        self.deskXL22.setItem(rowPosition, ck, item)

                        sn = 0
                        sb = 0
                        sm = 0


                        nstr = nstr + 1
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)

                    else:
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)
                else:
                    k_sss = k_sss + 1
                    sss = sss[0] + str(k_sss)
            
            if(self.k == 1):
                self.deskXL22.setGeometry(100, 190, 1670, self.hh)
            else:
                if(self.k == 2):
                    self.deskXL22.setGeometry(100, 190 + 10 + self.hh, 1670, self.hh)
                else:
                    self.deskXL22.setGeometry(100, 190 + 20 + 2 * self.hh, 1670, self.hh)
            self.deskXL22.show()
        
        self.bar.show()
        self.verSum.show()
        self.bar2.show()
        self.horSum.show()

    def ttt2(self):
        nstr = 0

        if(self.arrKnopok[0].isChecked() == True):
            sss = 'B39'
            k_sss = 39

            ob = self.ss20[sss].value    
            if(not(ob is None)):
                ob = ob.strip(' ') 
            while(ob != 'ВЫРУЧКА КТ'):
                ob = self.ss20[sss].value    
                if(not(ob is None)):
                    ob = ob.strip(' ') 
                    if(ob != 'КОЛИЧЕСТВО КТ' and ob != 'ВЫРУЧКА КТ'):
                        
                        rowPosition = self.deskXL20.rowCount()
                        self.deskXL20.insertRow(rowPosition)

                        item = QTableWidgetItem(ob)   
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFont(QFont('Corbel', 12))
                        self.deskXL20.setVerticalHeaderItem(rowPosition, item)
                        ck = column_index_from_string('D')
                        ck1 = get_column_letter(ck)
                        ckstr = ck1 + str(k_sss)
                        summ = 0
                        sn = 0
                        sm = 0
                        sb = 0

                        while(ck < column_index_from_string('AZ')):

                            item = self.ss20[ckstr].value
                            if(item is None):
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                                summ = summ + 0
                            else:
                                if(ck % 4 != 3):
                                    summ = summ + item
                                    if(ck % 4 == 0):
                                        sn = sn + item
                                    if(ck % 4 == 1):
                                        sm = sm + item
                                    if(ck % 4 == 2):
                                        sb = sb + item
                                    
                                else:
                                    item = summ
                                    summ = 0
                                
                                item = str(item)
                                item = QTableWidgetItem(item)
                                self.deskXL20.setItem(rowPosition, ck - 3, QTableWidgetItem(item))
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                            ckstr = ck1 + str(k_sss)

                        item = QTableWidgetItem(str(sn))
                        self.deskXL20.setItem(rowPosition, ck - 3, item)
                        item = QTableWidgetItem(str(sm))
                        self.deskXL20.setItem(rowPosition, ck - 2, item)
                        item = QTableWidgetItem(str(sb))
                        self.deskXL20.setItem(rowPosition, ck - 1, item)
                        item = QTableWidgetItem(str(sn + sb + sm))
                        self.deskXL20.setItem(rowPosition, ck, item)

                        sn = 0
                        sb = 0
                        sm = 0


                        nstr = nstr + 1
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)

                    elif(ob == 'КОЛИЧЕСТВО КТ'):
                        rowPosition = self.deskXL20.rowCount()
                        self.deskXL20.insertRow(rowPosition)
                        item = QTableWidgetItem(ob)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFont(QFont('Corbel', 12))
                        self.deskXL20.setVerticalHeaderItem(rowPosition, item)
                        ck = column_index_from_string('G')
                        ck1 = get_column_letter(ck)
                        ckstr = ck1 + str(k_sss)
                        summ = 0

                        while(ck < column_index_from_string('AZ')):

                            item = self.ss20[ckstr].value

                            if(item is None):
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                                summ = summ + 0
                            else:
                                if(ck % 4 == 3):
                                    summ = summ + int(item)
                                
                                    item = str(item)
                                    item = QTableWidgetItem(item)
                                    self.deskXL20.setItem(rowPosition, ck - 3, QTableWidgetItem(item))
                                    ck = ck + 1
                                    ck1 = get_column_letter(ck)

                            ckstr = ck1 + str(k_sss)

                        item = QTableWidgetItem(str(summ))
                        self.deskXL20.setItem(rowPosition, ck, item)

                        summ = 0

                        nstr = nstr + 1
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)
                    else:
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)
                else:
                    k_sss = k_sss + 1
                    sss = sss[0] + str(k_sss)

                summ3 = 0
                if(ob == 'ВЫРУЧКА КТ'):
                    rowPosition = self.deskXL20.rowCount()
                    self.deskXL20.insertRow(rowPosition)      
                    item = QTableWidgetItem(ob)
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFont(QFont('Corbel', 12))
                    self.deskXL20.setVerticalHeaderItem(rowPosition, item)
                    ck = column_index_from_string('G')
                    ck1 = get_column_letter(ck)
                    ckstr = ck1 + str(k_sss)
                    summ = 0

                    while(ck < column_index_from_string('AZ')):
                        summ2 = 0
                        for i in range (39, 45):
                            strrr = get_column_letter(ck - 3) + str(i)
                            obbb = self.ss20[strrr].value
                            #print(obbb)
                            if(not(obbb is None)):
                                summ2 = summ2 + int(obbb)

                            strrr = get_column_letter(ck - 2) + str(i)
                            obbb = self.ss20[strrr].value
                            #print(obbb)
                            if(not(obbb is None)):
                                summ2 = summ2 + int(obbb)

                            strrr = get_column_letter(ck - 1) + str(i)
                            obbb = self.ss20[strrr].value
                            #print(obbb)
                            #print(i)
                            if(not(obbb is None)):
                                summ2 = summ2 + int(obbb)

                        item = summ2
                        summ3 = summ3 + summ2
                            

                        if(item is None):
                            ck = ck + 4
                            ck1 = get_column_letter(ck)
                        else:
                            item = str(item)
                            item = QTableWidgetItem(item)
                            self.deskXL20.setItem(rowPosition, ck - 3, QTableWidgetItem(item))
                            ck = ck + 4
                            ck1 = get_column_letter(ck)

                        ckstr = ck1 + str(k_sss)

                    item = QTableWidgetItem(str(summ3))
                    self.deskXL20.setItem(rowPosition, ck - 3, item)

                    nstr = nstr + 1
                    k_sss = k_sss + 1
                    sss = sss[0] + str(k_sss)  

            self.deskXL20.setGeometry(100, 190, 1670, self.hh)
            self.deskXL20.show()
     
        if(self.arrKnopok[1].isChecked() == True):
            sss = 'B39'
            k_sss = 39

            ob = self.ss21[sss].value    
            if(not(ob is None)):
                ob = ob.strip(' ') 
            while(ob != 'ВЫРУЧКА КТ'):
                ob = self.ss21[sss].value    
                if(not(ob is None)):
                    ob = ob.strip(' ') 
                    if(ob != 'КОЛИЧЕСТВО КТ' and ob != 'ВЫРУЧКА КТ'):

                        rowPosition = self.deskXL21.rowCount()
                        self.deskXL21.insertRow(rowPosition)

                        item = QTableWidgetItem(ob)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFont(QFont('Corbel', 12))
                        self.deskXL21.setVerticalHeaderItem(rowPosition, item)
                        ck = column_index_from_string('D')
                        ck1 = get_column_letter(ck)
                        ckstr = ck1 + str(k_sss)
                        summ = 0
                        sn = 0
                        sm = 0
                        sb = 0

                        while(ck < column_index_from_string('AZ')):

                            item = self.ss21[ckstr].value
                            if(item is None):
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                                summ = summ + 0
                            else:
                                if(ck % 4 != 3):
                                    summ = summ + item
                                    if(ck % 4 == 0):
                                        sn = sn + item
                                    if(ck % 4 == 1):
                                        sm = sm + item
                                    if(ck % 4 == 2):
                                        sb = sb + item
                                    
                                else:
                                    item = summ
                                    summ = 0
                                
                                item = str(item)
                                item = QTableWidgetItem(item)
                                self.deskXL21.setItem(rowPosition, ck - 3, QTableWidgetItem(item))
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                            ckstr = ck1 + str(k_sss)

                        item = QTableWidgetItem(str(sn))
                        self.deskXL21.setItem(rowPosition, ck - 3, item)
                        item = QTableWidgetItem(str(sm))
                        self.deskXL21.setItem(rowPosition, ck - 2, item)
                        item = QTableWidgetItem(str(sb))
                        self.deskXL21.setItem(rowPosition, ck - 1, item)
                        item = QTableWidgetItem(str(sn + sb + sm))
                        self.deskXL21.setItem(rowPosition, ck, item)

                        sn = 0
                        sb = 0
                        sm = 0


                        nstr = nstr + 1
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)

                    elif(ob == 'КОЛИЧЕСТВО КТ'):
                        
                        rowPosition = self.deskXL21.rowCount()
                        self.deskXL21.insertRow(rowPosition)
                        item = QTableWidgetItem(ob)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFont(QFont('Corbel', 12))
                        self.deskXL21.setVerticalHeaderItem(rowPosition, item)
                        ck = column_index_from_string('G')
                        ck1 = get_column_letter(ck)
                        ckstr = ck1 + str(k_sss)
                        summ = 0

                        while(ck < column_index_from_string('AZ')):

                            item = self.ss21[ckstr].value

                            if(item is None):
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                                summ = summ + 0
                            else:
                                if(ck % 4 == 3):
                                    summ = summ + int(item)
                                
                                    item = str(item)
                                    item = QTableWidgetItem(item)
                                    self.deskXL21.setItem(rowPosition, ck - 3, QTableWidgetItem(item))
                                    ck = ck + 1
                                    ck1 = get_column_letter(ck)

                            ckstr = ck1 + str(k_sss)

                        item = QTableWidgetItem(str(summ))
                        self.deskXL21.setItem(rowPosition, ck, item)

                        summ = 0

                        nstr = nstr + 1
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)
                    else:
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)

                else:
                    k_sss = k_sss + 1
                    sss = sss[0] + str(k_sss)

                summ3 = 0
                if(ob == 'ВЫРУЧКА КТ'):
                    rowPosition = self.deskXL21.rowCount()
                    self.deskXL21.insertRow(rowPosition)      
                    item = QTableWidgetItem(ob)
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFont(QFont('Corbel', 12))
                    self.deskXL21.setVerticalHeaderItem(rowPosition, item)
                    ck = column_index_from_string('G')
                    ck1 = get_column_letter(ck)
                    ckstr = ck1 + str(k_sss)
                    summ = 0

                    while(ck < column_index_from_string('AZ')):
                        summ2 = 0
                        for i in range (39, 45):
                            strrr = get_column_letter(ck - 3) + str(i)
                            obbb = self.ss21[strrr].value
                            if(not(obbb is None)):
                                summ2 = summ2 + int(obbb)

                            strrr = get_column_letter(ck - 2) + str(i)
                            obbb = self.ss21[strrr].value
                            
                            if(not(obbb is None)):
                                summ2 = summ2 + int(obbb)
                            strrr = get_column_letter(ck - 1) + str(i)
                            obbb = self.ss21[strrr].value
                            if(not(obbb is None)):
                                summ2 = summ2 + int(obbb)

                        item = summ2
                        summ3 = summ3 + summ2
                            

                        if(item is None):
                            ck = ck + 4
                            ck1 = get_column_letter(ck)
                        else:
                            item = str(item)
                            item = QTableWidgetItem(item)
                            self.deskXL21.setItem(rowPosition, ck - 3, QTableWidgetItem(item))
                            ck = ck + 4
                            ck1 = get_column_letter(ck)

                        ckstr = ck1 + str(k_sss)

                    item = QTableWidgetItem(str(summ3))
                    self.deskXL21.setItem(rowPosition, ck - 3, item)

                    nstr = nstr + 1
                    k_sss = k_sss + 1
                    sss = sss[0] + str(k_sss)  



            if(self.k == 1):
                self.deskXL21.setGeometry(100, 190, 1670, self.hh)
            else:
                if(self.arrKnopok[0].isChecked() == True):
                    self.deskXL21.setGeometry(100, 190 + 10 + self.hh, 1670, self.hh)
                else:
                    self.deskXL21.setGeometry(100, 190, 1670, self.hh)
            self.deskXL21.show()
        
        if(self.arrKnopok[2].isChecked() == True):
            sss = 'B39'
            k_sss = 39

            ob = self.ss22[sss].value    
            if(not(ob is None)):
                ob = ob.strip(' ') 
            while(ob != 'ВЫРУЧКА КТ'):
                ob = self.ss22[sss].value    
                if(not(ob is None)):
                    ob = ob.strip(' ') 
                    if(ob != 'КОЛИЧЕСТВО КТ' and ob != 'ВЫРУЧКА КТ'):

                        rowPosition = self.deskXL22.rowCount()
                        self.deskXL22.insertRow(rowPosition)

                        item = QTableWidgetItem(ob)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFont(QFont('Corbel', 12))
                        self.deskXL22.setVerticalHeaderItem(rowPosition, item)
                        ck = column_index_from_string('D')
                        ck1 = get_column_letter(ck)
                        ckstr = ck1 + str(k_sss)
                        summ = 0
                        sn = 0
                        sm = 0
                        sb = 0

                        while(ck < column_index_from_string('AZ')):

                            item = self.ss22[ckstr].value
                            if(item is None):
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                                summ = summ + 0
                            else:
                                if(ck % 4 != 3):
                                    summ = summ + item
                                    if(ck % 4 == 0):
                                        sn = sn + item
                                    if(ck % 4 == 1):
                                        sm = sm + item
                                    if(ck % 4 == 2):
                                        sb = sb + item
                                    
                                else:
                                    item = summ
                                    summ = 0
                                
                                item = str(item)
                                item = QTableWidgetItem(item)
                                self.deskXL22.setItem(rowPosition, ck - 3, QTableWidgetItem(item))
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                            ckstr = ck1 + str(k_sss)

                        item = QTableWidgetItem(str(sn))
                        self.deskXL22.setItem(rowPosition, ck - 3, item)
                        item = QTableWidgetItem(str(sm))
                        self.deskXL22.setItem(rowPosition, ck - 2, item)
                        item = QTableWidgetItem(str(sb))
                        self.deskXL22.setItem(rowPosition, ck - 1, item)
                        item = QTableWidgetItem(str(sn + sb + sm))
                        self.deskXL22.setItem(rowPosition, ck, item)

                        sn = 0
                        sb = 0
                        sm = 0


                        nstr = nstr + 1
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)

                    elif(ob == 'КОЛИЧЕСТВО КТ'):
                        
                        rowPosition = self.deskXL22.rowCount()
                        self.deskXL22.insertRow(rowPosition)
                        item = QTableWidgetItem(ob)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFont(QFont('Corbel', 12))
                        self.deskXL22.setVerticalHeaderItem(rowPosition, item)
                        ck = column_index_from_string('G')
                        ck1 = get_column_letter(ck)
                        ckstr = ck1 + str(k_sss)
                        summ = 0

                        while(ck < column_index_from_string('AZ')):

                            item = self.ss22[ckstr].value

                            if(item is None):
                                ck = ck + 1
                                ck1 = get_column_letter(ck)
                                summ = summ + 0
                            else:
                                if(ck % 4 == 3):
                                    summ = summ + int(item)
                                
                                    item = str(item)
                                    item = QTableWidgetItem(item)
                                    self.deskXL22.setItem(rowPosition, ck - 3, QTableWidgetItem(item))
                                    ck = ck + 1
                                    ck1 = get_column_letter(ck)

                            ckstr = ck1 + str(k_sss)

                        item = QTableWidgetItem(str(summ))
                        self.deskXL22.setItem(rowPosition, ck, item)

                        summ = 0

                        nstr = nstr + 1
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)
                    else:
                        k_sss = k_sss + 1
                        sss = sss[0] + str(k_sss)

                else:
                    k_sss = k_sss + 1
                    sss = sss[0] + str(k_sss)

                summ3 = 0
                if(ob == 'ВЫРУЧКА КТ'):
                    rowPosition = self.deskXL22.rowCount()
                    self.deskXL22.insertRow(rowPosition)      
                    item = QTableWidgetItem(ob)
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFont(QFont('Corbel', 12))
                    self.deskXL22.setVerticalHeaderItem(rowPosition, item)
                    ck = column_index_from_string('G')
                    ck1 = get_column_letter(ck)
                    ckstr = ck1 + str(k_sss)
                    summ = 0

                    while(ck < column_index_from_string('AZ')):
                        summ2 = 0
                        for i in range (39, 45):
                            strrr = get_column_letter(ck - 3) + str(i)
                            obbb = self.ss20[strrr].value
                            #print(obbb)
                            if(not(obbb is None)):
                                summ2 = summ2 + int(obbb)

                            strrr = get_column_letter(ck - 2) + str(i)
                            obbb = self.ss20[strrr].value
                            #print(obbb)
                            if(not(obbb is None)):
                                summ2 = summ2 + int(obbb)

                            strrr = get_column_letter(ck - 1) + str(i)
                            obbb = self.ss20[strrr].value
                            #print(obbb)
                            #print(i)
                            if(not(obbb is None)):
                                summ2 = summ2 + int(obbb)

                        item = summ2
                        summ3 = summ3 + summ2
                            

                        if(item is None):
                            ck = ck + 4
                            ck1 = get_column_letter(ck)
                        else:
                            item = str(item)
                            item = QTableWidgetItem(item)
                            self.deskXL22.setItem(rowPosition, ck - 3, QTableWidgetItem(item))
                            ck = ck + 4
                            ck1 = get_column_letter(ck)

                        ckstr = ck1 + str(k_sss)

                    item = QTableWidgetItem(str(summ3))
                    self.deskXL22.setItem(rowPosition, ck - 3, item)

                    nstr = nstr + 1
                    k_sss = k_sss + 1
                    sss = sss[0] + str(k_sss)  



            if(self.k == 1):
                self.deskXL22.setGeometry(100, 190, 1670, self.hh)
            else:
                if(self.k == 2):
                    self.deskXL22.setGeometry(100, 190 + 10 + self.hh, 1670, self.hh)
                else:
                    self.deskXL22.setGeometry(100, 190 + 20 + 2 * self.hh, 1670, self.hh)
            self.deskXL22.show()

        self.bar.show()
        self.verSum.show()
        self.bar2.show()
        self.horSum.show()

    def check(self):
        self.arr20path = []                                                                                                              
        self.arr21path = []                                                                                                             
        self.arr22path = []    

        self.arr20n =[]
        self.arr20n.append("jan20.xlsx")
        self.arr20n.append("feb20.xlsx")
        self.arr20n.append("mar20.xlsx")
        self.arr20n.append("apr20.xlsx")
        self.arr20n.append("may20.xlsx")
        self.arr20n.append("jun20.xlsx")
        self.arr20n.append("jul20.xlsx")
        self.arr20n.append("aug20.xlsx")
        self.arr20n.append("sep20.xlsx")
        self.arr20n.append("oct20.xlsx")
        self.arr20n.append("nov20.xlsx")
        self.arr20n.append("dec20.xlsx")
        for i in range (0, 12):
            strpath = 'DATA/EXC/archive/2020/' + str(self.arr20n[i])
            if(os.path.exists(strpath)):
                self.arr20path.append(strpath)
            else:
                self.arr20path.append("x")

        self.arr21n =[]
        self.arr21n.append("jan21.xlsx")
        self.arr21n.append("feb21.xlsx")
        self.arr21n.append("mar21.xlsx")
        self.arr21n.append("apr21.xlsx")
        self.arr21n.append("may21.xlsx")
        self.arr21n.append("jun21.xlsx")
        self.arr21n.append("jul21.xlsx")
        self.arr21n.append("aug21.xlsx")
        self.arr21n.append("sep21.xlsx")
        self.arr21n.append("oct21.xlsx")
        self.arr21n.append("nov21.xlsx")
        self.arr21n.append("dec21.xlsx")
        for i in range (0, 12):
            strpath = 'DATA/EXC/archive/2021/' + str(self.arr21n[i])
            if(os.path.exists(strpath)):
                self.arr21path.append(strpath)
            else:
                self.arr21path.append("x")
        
        self.arr22n =[]
        self.arr22n.append("jan22.xlsx")
        self.arr22n.append("feb22.xlsx")
        self.arr22n.append("mar22.xlsx")
        self.arr22n.append("apr22.xlsx")
        self.arr22n.append("may22.xlsx")
        self.arr22n.append("jun22.xlsx")
        self.arr22n.append("jul22.xlsx")
        self.arr22n.append("aug22.xlsx")
        self.arr22n.append("sep22.xlsx")
        self.arr22n.append("oct22.xlsx")
        self.arr22n.append("nov22.xlsx")
        self.arr22n.append("dec22.xlsx")
        for i in range (0, 12):
            strpath = 'DATA/EXC/archive/2022/' + str(self.arr22n[i])
            if(os.path.exists(strpath)):
                self.arr22path.append(strpath)
            else:
                self.arr22path.append("x")

    def clear(self):
        for i in range (0, len(self.objects)):
            self.objects[i].hide()

    def interpreter(self, strk):        
        if(strk == 'jan22.xlsx'):
            return 'Январь 2022'
        if(strk == 'feb22.xlsx'):
            return 'Февраль 2022'
        if(strk == 'mar22.xlsx'):
            return 'Март 2022'
        if(strk == 'apr22.xlsx'):
            return 'Апрель 2022'
        if(strk == 'may22.xlsx'):
            return 'Май 2022'
        if(strk == 'jun22.xlsx'):
            return 'Июнь 2022'
        if(strk == 'jul22.xlsx'):
            return 'Июль 2022'
        if(strk == 'aug22.xlsx'):
            return 'Август 2022'
        if(strk == 'sep22.xlsx'):
            return 'Сентябрь 2022'
        if(strk == 'oct22.xlsx'):
            return 'Октябрь 2022'
        if(strk == 'nov22.xlsx'):
            return 'Ноябрь 2022'
        if(strk == 'dec22.xlsx'):
            return 'Декабрь 2022'
        if(strk == 'all22.xlsx'):
            return 'Весь 2022'

        if(strk == 'jan20.xlsx'):
            return 'Январь 2020'
        if(strk == 'feb20.xlsx'):
            return 'Февраль 2020'
        if(strk == 'mar20.xlsx'):
            return 'Март 2020'
        if(strk == 'apr20.xlsx'):
            return 'Апрель 2020'
        if(strk == 'may20.xlsx'):
            return 'Май 2020'
        if(strk == 'jun20.xlsx'):
            return 'Июнь 2020'
        if(strk == 'jul20.xlsx'):
            return 'Июль 2020'
        if(strk == 'aug20.xlsx'):
            return 'Август 2020'
        if(strk == 'sep20.xlsx'):
            return 'Сентябрь 2020'
        if(strk == 'oct20.xlsx'):
            return 'Октябрь 2020'
        if(strk == 'nov20.xlsx'):
            return 'Ноябрь 2020'
        if(strk == 'dec20.xlsx'):
            return 'Декабрь 2020'
        if(strk == 'all20.xlsx'):
            return 'Весь 2020'

        if(strk == 'jan21.xlsx'):
            return 'Январь 2021'
        if(strk == 'feb21.xlsx'):
            return 'Февраль 2021'
        if(strk == 'mar21.xlsx'):
            return 'Март 2021'
        if(strk == 'apr21.xlsx'):
            return 'Апрель 2021'
        if(strk == 'may21.xlsx'):
            return 'Май 2021'
        if(strk == 'jun21.xlsx'):
            return 'Июнь 2021'
        if(strk == 'jul21.xlsx'):
            return 'Июль 2021'
        if(strk == 'aug21.xlsx'):
            return 'Август 2021'
        if(strk == 'sep21.xlsx'):
            return 'Сентябрь 2021'
        if(strk == 'oct21.xlsx'):
            return 'Октябрь 2021'
        if(strk == 'nov21.xlsx'):
            return 'Ноябрь 2021'
        if(strk == 'dec21.xlsx'):
            return 'Декабрь 2021'
        if(strk == 'all21.xlsx'):
            return 'Весь 2021'
                                               
MAX_WIDTH = get_monitors()[0].width
MAX_HEIGHT = get_monitors()[0].height

app = QApplication(sys.argv)

if(MAX_WIDTH > 1900 and MAX_HEIGHT > 1000):
    win = MainWindow()


win.show()
app.exec()

